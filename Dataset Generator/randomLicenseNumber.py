import random
import openpyxl

# Define the list of states
letter = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
numbers = '0123456789'
company  = ['MarutiSuzuki','Hyundai','Tata','Toyota','Kia','Skoda','MercedesBenz','Volkswagen','Honda','Renault']
color = ['Red','White','silver','Blue','Black','Brown']
states = ['TN','AP','KA','KL','UP']


MarutiSuzuki = ['Desire', 'Ertiga']
Hyundai = ['Creta', 'Venue']
Tata = ['Nexon', 'Punch']
Toyota = ['Glanza', 'Fortuner']
Kia = ['Sonet', 'Seltos']
Skoda = ['Octavia', 'Kushaq']
MercedesBenz = ['GLA', 'S-class']
Volkswagen = ['Tiguan', 'Virtus']
Honda = ['City', 'Amaze']
Renault = ['KWID', 'Triber']


# Generate 10,000 random values from the list of states
data = [random.choice(states) for i in range(100000)]
stateNum1 = [random.choice(numbers) for i in range(100000)] 
stateNum2 = [random.choice(numbers) for i in range(100000)]
rto1 = [random.choice(letter) for i in range(100000)]
rto2 = [random.choice(letter) for i in range(100000)]
num1 =  [random.choice(numbers) for i in range(100000)] 
num2 =  [random.choice(numbers) for i in range(100000)] 
num3 =  [random.choice(numbers) for i in range(100000)] 
num4 =  [random.choice(numbers) for i in range(100000)] 
make =  [random.choice(company) for i in range(100000)] 
col = [random.choice(color) for i in range(100000)]

cars = []
for m in make:
    if m == 'MarutiSuzuki':
        model = random.choice(MarutiSuzuki)
    elif m == 'Hyundai':
        model = random.choice(Hyundai)
    elif m == 'Tata':
        model = random.choice(Tata)
    elif m == 'Toyota':
        model = random.choice(Toyota)
    elif m == 'Kia':
        model = random.choice(Kia)
    elif m == 'Skoda':
        model = random.choice(Skoda)
    elif m == 'MercedesBenz':
        model = random.choice(MercedesBenz)
    elif m == 'Volkswagen':
        model = random.choice(Volkswagen)
    elif m == 'Honda':
        model = random.choice(Honda)
    elif m == 'Renault':
        model = random.choice(Renault)
    cars.append({'make': m, 'model': model})


# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers
worksheet['A1'] = 'State'
worksheet['B1'] = 'District Code'
worksheet['C1'] = 'RTO'
worksheet['D1'] = 'UID'
worksheet['E1'] = 'Make'
worksheet['F1'] = 'Color'
worksheet['G1'] = 'Model'


# Write the data to the worksheet
for row in range(1, len(data) + 1):
    worksheet.cell(row=row+1, column=1, value=data[row-1])

for row in range(1, len(stateNum1) + 1):
    worksheet.cell(row=row+1, column=2, value=stateNum1[row-1]+stateNum2[row-1])

for row in range(1, len(rto1) + 1):
    worksheet.cell(row=row+1, column=3, value=rto1[row-1]+rto2[row-1])

for row in range(1, len(num1) + 1):
    worksheet.cell(row=row+1, column=4, value=num1[row-1]+num2[row-1]+num3[row-1]+num4[row-1])

for row in range(1, len(make) + 1):
    worksheet.cell(row=row+1, column=5, value=make[row-1])

for row in range(1, len(col) + 1):
    worksheet.cell(row=row+1, column=6, value=col[row-1])

for i, car in enumerate(cars):
    row = i + 2
    worksheet.cell(row=row, column=5, value=car['make'])
    worksheet.cell(row=row, column=7, value=car['model'])


# if(worksheet.cell == MarutiSuzuki):
#     model = [random.choice(MarutiSuzuki) for i in range(100000)]
#     for row in range(1, len(MarutiSuzuki) + 1):
#         worksheet.cell(row=row+1, column=8, value=model[row-1])

# Save the workbook
workbook.save('random_data.xlsx')



